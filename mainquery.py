import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
import io
import msoffcrypto # type: ignore


def unlock_excel(file_content, password):
    try:
        # Use msoffcrypto to handle the password protection
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(file_content)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

        # Load the decrypted file into openpyxl
        decrypted.seek(0)
        workbook = load_workbook(decrypted, read_only=False, data_only=True, keep_vba=False, keep_links=False)

        # Save the unlocked workbook to a BytesIO object
        unlocked_file = io.BytesIO()
        workbook.save(unlocked_file)
        unlocked_file.seek(0)
        return unlocked_file
    except Exception as e:
        raise Exception(f"{str(e)}")

def load_template_headers():
    try:
        # Load the template headers from the "Template Header" column of the external Excel file
        mappings_df = pd.read_excel('header_mappings.xlsx')
        if 'Template Header' not in mappings_df.columns:
            st.error("The header_mappings.xlsx file must contain 'Template Header' column.")
            return None
        
        template_headers = mappings_df['Template Header'].dropna().unique().tolist()
        template_headers = [header.strip().upper() for header in template_headers]
        return template_headers
    except Exception as e:
        st.error(f"Error loading template headers: {e}")
        return None
    
# Load template headers from the Excel file
template_headers = load_template_headers()
if template_headers is None:
    st.stop()

def load_header_mappings():
    try:
        # Load the header mappings from an external Excel file
        mappings_df = pd.read_excel('header_mappings.xlsx')
        
        # Verify that the required columns exist
        if 'Template Header' not in mappings_df.columns or 'Possible Headers' not in mappings_df.columns:
            st.error("The header_mappings.xlsx file must contain 'Template Header' and 'Possible Headers' columns.")
            return None
        
        header_mappings = {}
        for _, row in mappings_df.iterrows():
            template_header = str(row['Template Header']).strip().upper() if not pd.isna(row['Template Header']) else ''
            possible_headers = str(row['Possible Headers']).strip().upper().split(', ') if not pd.isna(row['Possible Headers']) else []
            if template_header not in header_mappings:
                header_mappings[template_header] = []
            header_mappings[template_header].extend(possible_headers)
        return header_mappings
    
    except Exception as e:
        st.error(f"Error loading header mappings: {e}")
        return None

def map_headers(df, header_mappings):
    if header_mappings is None:
        return {}
    
    mapping = {}
    uploaded_headers_upper = {header.upper().strip(): header for header in df.columns}
    for template_header in template_headers:
        template_header_upper = template_header.upper()
        possible_headers = header_mappings.get(template_header_upper, [])
        mapped_header = None
        for possible_header in possible_headers:
            if possible_header in uploaded_headers_upper:
                mapped_header = uploaded_headers_upper[possible_header]
                break
        mapping[template_header] = mapped_header
   
    return mapping
    
def add_ch_code_prefix(df):
    # Check for 'CH CODE' column (case-insensitive)
    ch_code_col = next((col for col in df.columns if col.strip().upper() == 'CH CODE'), None)
    
    if ch_code_col is None:
        print("CH CODE column not found in the uploaded file.")
        return df

    # Ensure 'PREFIX' column exists, initialize with None if it doesn't
    if 'PREFIX' not in df.columns:
        df['PREFIX'] = None

    # Process each row to update 'PREFIX' column
    for index, row in df.iterrows():
        ch_code = row[ch_code_col]
        if pd.notnull(ch_code) and isinstance(ch_code, str):
            # Remove hyphens and take the first 6 characters
            cleaned_code = ch_code.replace('-', '')
            prefix = cleaned_code[:6]  # Extract first 6 characters as prefix
            df.at[index, 'PREFIX'] = prefix  # Update the 'PREFIX' column in the DataFrame
        else:
            df.at[index, 'PREFIX'] = None  # Handle NaN or NaT values

    return df


def process_each_sheet(uploaded_file):
    try:
        # Load the uploaded Excel file
        xls = pd.ExcelFile(uploaded_file)
        dfs = []
        sheet_names = []

        for sheet_name in xls.sheet_names:
            # Read each sheet into a DataFrame
            df = pd.read_excel(xls, sheet_name=sheet_name)
            dfs.append(df)
            sheet_names.append(sheet_name)

        return dfs, sheet_names
    except Exception as e:
        st.error(f"Error processing Excel sheets: {e}")
        return None, None
    
# def fill_missing_headers(df, template_headers):
#     for header in template_headers:
#         if header not in df.columns:
#             df[header] = None
#     return df

# def compare_dataframes(final_df, template_df):
#     mismatched_rows = []
#     for index, row in final_df.iterrows():
#         if not row.equals(template_df.loc[index]):
#             mismatched_rows.append(index)
#     return mismatched_rows

def main():
    st.title("Automation Selective tool")
    st.markdown("""
    ### Instructions:            
    1. Consider removing the design of Raw file if the first (1) COL and ROW is not HEADER
    2. Before the automation please FEED the possible headers to make it accurate
    3. Upload CSV/XLSX file
    4. Input the CSV/XLSX password if the system detects it as encrypted.
    5. Download the OUTPUT file.
    6. Expect it may be slow due to QUERY from BCRM.

    (Note: It does not accept xls file, consider resave the file as csv or xlsx)    
    """)

    # Step 1: Upload the file
    uploaded_file = st.file_uploader("Upload an Excel or CSV file", type=["xlsx", "csv"])

    if uploaded_file is not None:
        # Extract the original file name without extension
        original_file_name = os.path.splitext(uploaded_file.name)[0]
        
        # Step 2: Check if file is an Excel file and prompt for password if encrypted
        if uploaded_file.name.endswith('.xlsx'):
            file_content = io.BytesIO(uploaded_file.getvalue())
            try:
                # Try to load the workbook without a password
                workbook = load_workbook(file_content, read_only=False, data_only=True, keep_vba=False, keep_links=False)
                st.toast("File is not password protected, proceeding with automation! ✔️")
                dfs, sheet_names = process_each_sheet(uploaded_file)
            except Exception:
                password = st.text_input("Password", type="password")
                if password:
                    try:
                        unlocked_file = unlock_excel(file_content, password)
                        st.toast("File unlocked successfully! ✔️")
                        with st.spinner("Processing selectives..."):
                            try:
                                dfs, sheet_names = process_each_sheet(unlocked_file)
                                st.write("Excel sheets processed successfully.")
                            except Exception as e:
                                st.error(f"Failed to process the unlocked file: {e}")
                                st.stop()
                    except Exception as e:
                        st.error(str(e))    
                        st.stop()
                else:
                    st.info("Please provide a password to unlock the file.")
                    st.stop()
        elif uploaded_file.name.endswith('.csv'):
            with st.spinner("Processing selectives..."):
                try:
                    df = pd.read_csv(uploaded_file, encoding='utf-8')
                    dfs = [df]
                    st.write("CSV file processed successfully.")
                except UnicodeDecodeError:
                    st.error(f"Failed to read the CSV file with 'utf-8' encoding. Trying 'latin1' encoding. {uploaded_file.name}")
                    try:
                        df = pd.read_csv(uploaded_file, encoding='latin1')
                        dfs = [df]
                        st.write("CSV file processed successfully with 'latin1' encoding.")
                    except UnicodeDecodeError:
                        st.error(f"Failed to read the CSV file with 'latin1' encoding. Trying 'iso-8859-1' encoding. {uploaded_file.name}")
                        try:
                            df = pd.read_csv(uploaded_file, encoding='iso-8859-1')
                            dfs = [df]
                            st.write("CSV file processed successfully with 'iso-8859-1' encoding.")
                        except UnicodeDecodeError:
                            st.error(f"Failed to read the CSV file with 'iso-8859-1' encoding. {uploaded_file.name}")
                            st.stop()
        
        else:
            st.info("Please upload a file.")
            st.stop()

        with st.status("Processing Selective...", expanded=False) as status:
            # Convert all column names to strings to avoid mixed-type warnings
            for df in dfs:
                df.columns = df.columns.map(str)
            
            st.write("Uploaded file preview:")
            st.dataframe(dfs[0].head())
            
            # Load header mappings from the database file
            header_mappings = load_header_mappings()
            if header_mappings is None:
                st.stop()

            # Prepare final DataFrame to hold appended data
            final_df = pd.DataFrame(columns=template_headers)
            
            for df, sheet_name in zip(dfs, sheet_names):
                
                # Step 5: Map the headers
                mapping = map_headers(df, header_mappings)
                
                # Check if the sheet contains at least two template headers
                valid_headers = [header for header in mapping.values() if header is not None]
                if len(valid_headers) < 2:
                    st.info(f"Skipping sheet ({sheet_name.upper()}) as it does not contain at least two template headers.")
                    continue
                st.write(sheet_name)
                st.json(mapping, expanded=False)

                # Step 6: Extract values based on mapped headers
                extracted_data = {}
                for template_header, original_header in mapping.items():
                    if original_header is not None:
                        extracted_data[template_header] = df[original_header].copy()
                    else:
                        extracted_data[template_header] = pd.Series([None] * len(df))


                        mapped_df = pd.DataFrame(extracted_data)
                final_df = pd.concat([final_df, mapped_df], ignore_index=True)
            # Step 3: Add CH CODE Prefix to First Column if CH CODE exists
            final_df = add_ch_code_prefix(final_df)
            
            st.write("OUTPUT PREVIEW:")
            st.dataframe(final_df.head())

            # final_df = fill_missing_headers(final_df, template_header)

            # template_header = query_database()

            # mismatched_rows = compare_dataframes(final_df, template_df)
            # if mismatched_rows:
            #     st.write(f"Mismatched: {mismatched_rows}")
            # else:
            #     st.write("All rows match the expected template.")

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Sheet1')

                # Auto-fit column widths
                worksheet = writer.sheets['Sheet1']
                for col in worksheet.columns:
                    max_length = 0
                    column_letter = col[0].column_letter  # Get the column letter
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if isinstance(cell.value, (int, float)):
                                cell_value_str = f"{cell.value:.0f}"  # Format as integer without decimal places
                            else:
                                cell_value_str = str(cell.value)
                            
                            if len(cell_value_str) > max_length:
                                max_length = len(cell_value_str)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)

            # Step 8: Provide download link for the templated Excel file
            current_date = datetime.datetime.now().strftime("%Y%m%d")
            new_file_name = f"{current_date}-{original_file_name}.xlsx"

            status.update(label=f"Process Complete!. {uploaded_file.name}", expanded=False)
        pass

        st.download_button(
            label="Download Mapped Excel",
            data=output,
            file_name=new_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    main()
